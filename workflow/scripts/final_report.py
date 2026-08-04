from pathlib import Path
import re
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RESULTS = Path("results")

SAMPLES = list(snakemake.params.samples)
SPECIES_GROUPS = dict(snakemake.params.species_groups)

SPECIES_NAMES = {
    "abaumannii": "Acinetobacter baumannii",
    "paeruginosa": "Pseudomonas aeruginosa",
    "kpneumoniae": "Klebsiella pneumoniae",
    "ecoli": "Escherichia coli",
}

SPECIES = {}

for species_key, samples in snakemake.params.species_groups.items():
    species_name = SPECIES_NAMES.get(species_key, species_key)

    for sample in samples:
        SPECIES[sample] = species_name


def parse_assembly_stats(sample):
    path = RESULTS / sample / "assembly_stats" / f"{sample}.txt"

    data = {
        "Genome size (bp)": None,
        "Scaffolds": None,
        "Largest scaffold (bp)": None,
        "N50 (bp)": None,
        "Gaps": None,
    }

    text = path.read_text()

    match = re.search(
        r"sum = (\d+), n = (\d+), ave = [\d.]+, largest = (\d+)",
        text
    )

    if match:
        data["Genome size (bp)"] = int(match.group(1))
        data["Scaffolds"] = int(match.group(2))
        data["Largest scaffold (bp)"] = int(match.group(3))

    match = re.search(r"N50 = (\d+)", text)

    if match:
        data["N50 (bp)"] = int(match.group(1))

    match = re.search(r"Gaps = (\d+)", text)

    if match:
        data["Gaps"] = int(match.group(1))

    return data


def parse_busco(sample):
    path = RESULTS / sample / "busco" / "busco_summary.txt"

    data = {
        "BUSCO Complete (%)": None,
        "BUSCO Single (%)": None,
        "BUSCO Duplicated (%)": None,
        "BUSCO Fragmented (%)": None,
        "BUSCO Missing (%)": None,
    }

    text = path.read_text()

    match = re.search(
        r"C:([\d.]+)%\[S:([\d.]+)%,D:([\d.]+)%\],F:([\d.]+)%,M:([\d.]+)%",
        text
    )

    if match:
        data["BUSCO Complete (%)"] = float(match.group(1))
        data["BUSCO Single (%)"] = float(match.group(2))
        data["BUSCO Duplicated (%)"] = float(match.group(3))
        data["BUSCO Fragmented (%)"] = float(match.group(4))
        data["BUSCO Missing (%)"] = float(match.group(5))

    return data


def parse_mlst(sample):
    path = RESULTS / sample / "mlst" / f"{sample}.tsv"

    fields = path.read_text().strip().split("\t")

    return {
        "MLST scheme": fields[1] if len(fields) > 1 else None,
        "ST": fields[2] if len(fields) > 2 else None,
    }


def parse_amrfinder(sample):
    path = RESULTS / sample / "amrfinder" / "amrfinder.tsv"

    genes = []

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:
            symbol = row.get("Element symbol")

            if symbol and symbol != "NA":
                genes.append(symbol)

    return sorted(set(genes))


def parse_mobsuite(sample):
    path = RESULTS / sample / "mobsuite" / "mobtyper_results.txt"

    if not path.exists():
        return []

    plasmids = []

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:
            plasmids.append(row)

    return plasmids

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            35
        )

    ws.row_dimensions[1].height = 30

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

wb = Workbook()

ws = wb.active
ws.title = "Summary"

headers = [
    "Sample",
    "Species",
    "Genome size (bp)",
    "Scaffolds",
    "Largest scaffold (bp)",
    "N50 (bp)",
    "Gaps",
    "BUSCO Complete (%)",
    "BUSCO Single (%)",
    "BUSCO Duplicated (%)",
    "BUSCO Fragmented (%)",
    "BUSCO Missing (%)",
    "MLST scheme",
    "ST",
    "AMR determinants",
    "Predicted plasmids",
]

ws.append(headers)

for sample in SAMPLES:

    assembly = parse_assembly_stats(sample)
    busco = parse_busco(sample)
    mlst = parse_mlst(sample)
    amr = parse_amrfinder(sample)
    plasmids = parse_mobsuite(sample)

    ws.append([
        sample,
        SPECIES[sample],
        assembly["Genome size (bp)"],
        assembly["Scaffolds"],
        assembly["Largest scaffold (bp)"],
        assembly["N50 (bp)"],
        assembly["Gaps"],
        busco["BUSCO Complete (%)"],
        busco["BUSCO Single (%)"],
        busco["BUSCO Duplicated (%)"],
        busco["BUSCO Fragmented (%)"],
        busco["BUSCO Missing (%)"],
        mlst["MLST scheme"],
        mlst["ST"],
        ", ".join(amr),
        len(plasmids),
    ])


for cell in ws[1]:
    cell.font = Font(bold=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

auto_width(ws)

# ============================================================
# AMR DETAIL
# ============================================================

ws_amr = wb.create_sheet("AMR")

amr_headers = [
    "Sample",
    "Species",
    "Contig",
    "Start",
    "Stop",
    "Strand",
    "Element symbol",
    "Element name",
    "Scope",
    "Type",
    "Subtype",
    "Class",
    "Subclass",
    "Method",
    "% Coverage",
    "% Identity",
    "Closest reference accession",
    "Closest reference name",
]

ws_amr.append(amr_headers)

for sample in SAMPLES:

    path = RESULTS / sample / "amrfinder" / "amrfinder.tsv"

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:

            ws_amr.append([
                sample,
                SPECIES[sample],
                row.get("Contig id"),
                row.get("Start"),
                row.get("Stop"),
                row.get("Strand"),
                row.get("Element symbol"),
                row.get("Element name"),
                row.get("Scope"),
                row.get("Type"),
                row.get("Subtype"),
                row.get("Class"),
                row.get("Subclass"),
                row.get("Method"),
                row.get("% Coverage of reference"),
                row.get("% Identity to reference"),
                row.get("Closest reference accession"),
                row.get("Closest reference name"),
            ])


for cell in ws_amr[1]:
    cell.font = Font(bold=True)

ws_amr.freeze_panes = "A2"
ws_amr.auto_filter.ref = ws_amr.dimensions

auto_width(ws_amr)

# ============================================================
# VFDB / VIRULENCE DETAIL
# ============================================================

ws_vfdb = wb.create_sheet("Virulence")

vfdb_headers = [
    "Sample",
    "Species",
    "Query protein",
    "VFDB hit",
    "% Identity",
    "Alignment length",
    "Mismatches",
    "Gap opens",
    "Query start",
    "Query end",
    "Subject start",
    "Subject end",
    "E-value",
    "Bit score",
    "% Query coverage",
]

ws_vfdb.append(vfdb_headers)

for sample in SAMPLES:

    path = RESULTS / sample / "vfdb" / f"{sample}_vfdb.tsv"

    with path.open() as handle:

        reader = csv.reader(handle, delimiter="\t")

        for row in reader:

            if len(row) < 13:
                continue

            ws_vfdb.append([
                sample,
                SPECIES[sample],
                row[0],
                row[1],
                float(row[2]),
                int(row[3]),
                int(row[4]),
                int(row[5]),
                int(row[6]),
                int(row[7]),
                int(row[8]),
                int(row[9]),
                row[10],
                float(row[11]),
                float(row[12]),
            ])


for cell in ws_vfdb[1]:
    cell.font = Font(bold=True)

ws_vfdb.freeze_panes = "A2"
ws_vfdb.auto_filter.ref = ws_vfdb.dimensions

auto_width(ws_vfdb)

# ============================================================
# MOB-SUITE / PLASMIDS
# ============================================================

ws_mob = wb.create_sheet("Plasmids")

mob_headers = [
    "Sample",
    "Species",
    "Plasmid ID",
    "Number of contigs",
    "Size (bp)",
    "GC",
    "Replicon type",
    "Relaxase type",
    "MPF type",
    "oriT type",
    "Predicted mobility",
    "Mash nearest neighbor",
    "Mash distance",
    "Mash identification",
    "Primary cluster",
    "Secondary cluster",
    "Predicted host range",
]

ws_mob.append(mob_headers)

for sample in SAMPLES:

    path = RESULTS / sample / "mobsuite" / "mobtyper_results.txt"

    if not path.exists():
        continue

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:
            ws_mob.append([
                sample,
                SPECIES[sample],
                row.get("sample_id"),
                row.get("num_contigs"),
                row.get("size"),
                row.get("gc"),
                row.get("rep_type(s)"),
                row.get("relaxase_type(s)"),
                row.get("mpf_type"),
                row.get("orit_type(s)"),
                row.get("predicted_mobility"),
                row.get("mash_nearest_neighbor"),
                row.get("mash_neighbor_distance"),
                row.get("mash_neighbor_identification"),
                row.get("primary_cluster_id"),
                row.get("secondary_cluster_id"),
                row.get("predicted_host_range_overall_name"),
            ])


for cell in ws_mob[1]:
    cell.font = Font(bold=True)

ws_mob.freeze_panes = "A2"
ws_mob.auto_filter.ref = ws_mob.dimensions

auto_width(ws_mob)

# ============================================================
# FASTANI
# ============================================================

ws_ani = wb.create_sheet("FastANI")

ani_headers = [
    "Query",
    "Reference",
    "ANI (%)",
    "Mapped fragments",
    "Total fragments",
]

ws_ani.append(ani_headers)

fastani_dir = RESULTS / "fastani"

if fastani_dir.exists():

    for path in fastani_dir.glob("*/fastani.tsv"):

        with path.open() as handle:

            reader = csv.reader(handle, delimiter="\t")

            for row in reader:

                if len(row) < 5:
                    continue

                query = Path(row[0]).parts[1]
                reference = Path(row[1]).parts[1]

                ws_ani.append([
                    query,
                    reference,
                    float(row[2]),
                    int(row[3]),
                    int(row[4]),
                ])


for cell in ws_ani[1]:
    cell.font = Font(bold=True)

ws_ani.freeze_panes = "A2"
ws_ani.auto_filter.ref = ws_ani.dimensions

auto_width(ws_ani)
# ============================================================
# IQ-TREE / PHYLOGENY
# ============================================================

ws_tree = wb.create_sheet("Phylogeny")

ws_tree.append([
    "Species",
    "Parameter",
    "Value",
])

iqtree_dir = RESULTS / "iqtree"

if iqtree_dir.exists():

    for species_dir in sorted(iqtree_dir.iterdir()):

        if not species_dir.is_dir():
            continue

        species = species_dir.name

        treefile = species_dir / "core_genome.treefile"
        iqtree_report = species_dir / "core_genome.iqtree"

        # Newick tree
        if treefile.exists():

            newick = treefile.read_text().strip()

            ws_tree.append([
                species,
                "Maximum-likelihood tree (Newick)",
                newick
            ])

        # IQ-TREE report
        if iqtree_report.exists():

            text = iqtree_report.read_text()

            model = re.search(
                r"Best-fit model according to BIC:\s+(\S+)",
                text
            )

            if not model:
                model = re.search(
                    r"Best-fit model:\s+(\S+)",
                    text
                )

            if model:
                ws_tree.append([
                    species,
                    "Best-fit substitution model",
                    model.group(1)
                ])

            likelihood = re.search(
                r"Log-likelihood of the tree:\s+([-\d.]+)",
                text
            )

            if likelihood:
                ws_tree.append([
                    species,
                    "Log-likelihood",
                    float(likelihood.group(1))
                ])

            tree_length = re.search(
                r"Sum of branch lengths:\s+([\d.eE+-]+)",
                text
            )

            if tree_length:
                ws_tree.append([
                    species,
                    "Total tree length",
                    float(tree_length.group(1))
                ])

        # Bootstrap only makes sense when enough sequences are present
        samples_in_species = SPECIES_GROUPS.get(species, [])

        if len(samples_in_species) < 4:
            ws_tree.append([
                species,
                "Bootstrap",
                "Not performed (fewer than 4 sequences)"
            ])


for cell in ws_tree[1]:
    cell.font = Font(bold=True)

ws_tree.freeze_panes = "A2"
ws_tree.auto_filter.ref = ws_tree.dimensions

ws_tree.column_dimensions["A"].width = 22
ws_tree.column_dimensions["B"].width = 35
ws_tree.column_dimensions["C"].width = 100

for row in ws_tree.iter_rows(min_row=2):
    row[2].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )
