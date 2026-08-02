from pathlib import Path
import re
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


RESULTS = Path("results")

SAMPLES = ["AB68", "AB69", "AB70", "AB71"]

SPECIES = {
    "AB68": "Acinetobacter baumannii",
    "AB69": "Pseudomonas aeruginosa",
    "AB70": "Acinetobacter baumannii",
    "AB71": "Acinetobacter baumannii",
}


def parse_assembly_stats(sample):
    path = RESULTS / sample / "assembly_stats" / f"{sample}.txt"

    data = {
        "Genome size (bp)": None,
        "Scaffolds": None,
        "Largest scaffold (bp)": None,
        "N50 (bp)": None,
        "Gaps": None,
    }

    text = path.read_text()

    match = re.search(
        r"sum = (\d+), n = (\d+), ave = [\d.]+, largest = (\d+)",
        text
    )

    if match:
        data["Genome size (bp)"] = int(match.group(1))
        data["Scaffolds"] = int(match.group(2))
        data["Largest scaffold (bp)"] = int(match.group(3))

    match = re.search(r"N50 = (\d+)", text)

    if match:
        data["N50 (bp)"] = int(match.group(1))

    match = re.search(r"Gaps = (\d+)", text)

    if match:
        data["Gaps"] = int(match.group(1))

    return data


def parse_busco(sample):
    path = RESULTS / sample / "busco" / "busco_summary.txt"

    data = {
        "BUSCO Complete (%)": None,
        "BUSCO Single (%)": None,
        "BUSCO Duplicated (%)": None,
        "BUSCO Fragmented (%)": None,
        "BUSCO Missing (%)": None,
    }

    text = path.read_text()

    match = re.search(
        r"C:([\d.]+)%\[S:([\d.]+)%,D:([\d.]+)%\],F:([\d.]+)%,M:([\d.]+)%",
        text
    )

    if match:
        data["BUSCO Complete (%)"] = float(match.group(1))
        data["BUSCO Single (%)"] = float(match.group(2))
        data["BUSCO Duplicated (%)"] = float(match.group(3))
        data["BUSCO Fragmented (%)"] = float(match.group(4))
        data["BUSCO Missing (%)"] = float(match.group(5))

    return data


def parse_mlst(sample):
    path = RESULTS / sample / "mlst" / f"{sample}.tsv"

    fields = path.read_text().strip().split("\t")

    return {
        "MLST scheme": fields[1] if len(fields) > 1 else None,
        "ST": fields[2] if len(fields) > 2 else None,
    }


def parse_amrfinder(sample):
    path = RESULTS / sample / "amrfinder" / "amrfinder.tsv"

    genes = []

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:
            symbol = row.get("Element symbol")

            if symbol and symbol != "NA":
                genes.append(symbol)

    return sorted(set(genes))


def parse_mobsuite(sample):
    path = RESULTS / "MOBsuite" / sample / "mobtyper_results.txt"

    if not path.exists():
        return []

    plasmids = []

    with path.open() as handle:
        reader = csv.DictReader(handle, delimiter="\t")

        for row in reader:
            plasmids.append(row)

    return plasmids
def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


wb = Workbook()

ws = wb.active
ws.title = "Summary"

headers = [
    "Sample",
    "Species",
    "Genome size (bp)",
    "Scaffolds",
    "Largest scaffold (bp)",
    "N50 (bp)",
    "Gaps",
    "BUSCO Complete (%)",
    "BUSCO Single (%)",
    "BUSCO Duplicated (%)",
    "BUSCO Fragmented (%)",
    "BUSCO Missing (%)",
    "MLST scheme",
    "ST",
    "AMR determinants",
    "Predicted plasmids",
]

ws.append(headers)

for sample in SAMPLES:

    assembly = parse_assembly_stats(sample)
    busco = parse_busco(sample)
    mlst = parse_mlst(sample)
    amr = parse_amrfinder(sample)
    plasmids = parse_mobsuite(sample)

    ws.append([
        sample,
        SPECIES[sample],
        assembly["Genome size (bp)"],
        assembly["Scaffolds"],
        assembly["Largest scaffold (bp)"],
        assembly["N50 (bp)"],
        assembly["Gaps"],
        busco["BUSCO Complete (%)"],
        busco["BUSCO Single (%)"],
        busco["BUSCO Duplicated (%)"],
        busco["BUSCO Fragmented (%)"],
        busco["BUSCO Missing (%)"],
        mlst["MLST scheme"],
        mlst["ST"],
        ", ".join(amr),
        len(plasmids),
    ])


for cell in ws[1]:
    cell.font = Font(bold=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

auto_width(ws)


output_dir = RESULTS / "final_report"
output_dir.mkdir(parents=True, exist_ok=True)

output = output_dir / "genomic_report.xlsx"

wb.save(output)

print(f"Report written to: {output}")
